from __future__ import annotations

import argparse
import csv
import getpass
import json
import os
import re
import sys
import uuid
from dataclasses import dataclass
from http.cookiejar import CookieJar
from typing import Any, Dict, Iterable, Optional
from urllib.error import HTTPError, URLError
from urllib.parse import urljoin
from urllib.request import HTTPCookieProcessor, Request, build_opener


@dataclass
class ApiClient:
    base_url: str

    def __post_init__(self) -> None:
        if not self.base_url.endswith("/"):
            self.base_url += "/"
        self.cookies = CookieJar()
        self.opener = build_opener(HTTPCookieProcessor(self.cookies))

    def request_json(
        self,
        method: str,
        path: str,
        *,
        body: Optional[Dict[str, Any]] = None,
        headers: Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        url = urljoin(self.base_url, path.lstrip("/"))
        data = None
        req_headers = {"accept": "application/json"}
        if headers:
            req_headers.update(headers)
        if body is not None:
            data = json.dumps(body).encode("utf-8")
            req_headers.setdefault("content-type", "application/json")
        req = Request(url=url, data=data, method=method.upper(), headers=req_headers)
        try:
            with self.opener.open(req, timeout=60) as resp:
                raw = resp.read().decode("utf-8")
                if not raw:
                    return {}
                return json.loads(raw)
        except HTTPError as e:
            raw = ""
            try:
                raw = e.read().decode("utf-8")
            except Exception:
                pass
            raise RuntimeError(f"HTTP {e.code} {e.reason} for {method} {url}: {raw}") from e
        except URLError as e:
            raise RuntimeError(f"Network error for {method} {url}: {e}") from e


def _default_customer_payload(name: str, code: str) -> Dict[str, Any]:
    return {
        "code": code,
        "name": name,
        "abn": None,
        "tax_id": None,
        "status": "Active",
        "contacts": [
            {
                "type": "Primary Contact",
                "name": "Test Admin",
                "title": "Admin",
                "email": "admin@example.com",
                "phone": "+61 400 000 000",
                "phone_alt": None,
                "preferred_method": "Email",
                "notes": "Seeded via plate import script",
            }
        ],
        "delivery_addresses": [
            {
                "label": "Head Office",
                "type": "Both",
                "street1": "1 Test Street",
                "street2": None,
                "suburb": "Sydney",
                "state": "NSW",
                "postcode": "2000",
                "country": "Australia",
                "contact_name": "Dispatch",
                "contact_phone": "+61 400 000 000",
                "delivery_instructions": "Leave at reception",
                "is_default": True,
            }
        ],
        "delivery_preferences": {
            "preferred_pallet_type": "Plain",
            "preferred_transport_company": None,
            "preferred_wrapping": True,
            "special_instructions": None,
            "delivery_contact_id": None,
        },
        "payment_terms": "Net 30",
        "credit_limit": 25000,
        "currency_preference": "AUD",
        "notes": "Seeded customer from plate database",
        "internal_notes": "Test data",
    }


def _generate_fallback_code() -> str:
    h = uuid.uuid4().hex[:4].upper()
    out = []
    for ch in h:
        if "0" <= ch <= "9":
            out.append(chr(ord("A") + int(ch)))
        else:
            out.append(chr(ord("K") + (ord(ch) - ord("A"))))
    return "".join(out)[:4]


def _derive_customer_code(name: str, used: set[str]) -> str:
    # CustomerCreateRequest expects ^[A-Z]{2,4}$
    base = re.sub(r"[^A-Za-z]+", " ", (name or "").strip()).strip()
    parts = [p for p in base.split(" ") if p]
    cand = ""
    if len(parts) >= 2:
        cand = (parts[0][:2] + parts[1][:2]).upper()
    elif len(parts) == 1:
        cand = (parts[0][:4]).upper()
    cand = re.sub(r"[^A-Z]", "", cand)
    cand = cand[:4]
    if len(cand) < 2:
        cand = _generate_fallback_code()

    code = cand
    if code in used:
        # deterministic-ish suffix: try to find a free variant by cycling letters
        for i in range(26):
            alt = (code[:3] + chr(ord("A") + i))[:4]
            if len(alt) >= 2 and alt not in used:
                code = alt
                break
        else:
            code = _generate_fallback_code()

    used.add(code)
    return code


def _norm_key(k: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (k or "").strip().lower())


def _get(row: dict[str, Any], keys: Iterable[str]) -> str:
    want = {_norm_key(k) for k in keys}
    for rk, rv in row.items():
        if _norm_key(rk) in want:
            return "" if rv is None else str(rv).strip()
    return ""


def _iter_rows(path: str, delimiter: str | None) -> Iterable[dict[str, Any]]:
    ext = os.path.splitext(path)[1].lower()
    if ext in [".csv", ".tsv", ".txt"]:
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            if delimiter is None:
                sample = f.read(4096)
                f.seek(0)
                sniff = csv.Sniffer()
                try:
                    dialect = sniff.sniff(sample, delimiters=[",", "\t", ";", "|"])
                except Exception:
                    dialect = csv.excel
                delim = getattr(dialect, "delimiter", ",")
            else:
                delim = delimiter
            reader = csv.DictReader(f, delimiter=delim)
            for r in reader:
                if not r:
                    continue
                yield r
        return

    if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        try:
            import openpyxl  # type: ignore
        except Exception as e:
            raise SystemExit(
                f"ERROR: {path} looks like an Excel file, but openpyxl isn't available. "
                f"Either export to CSV/TSV, or install openpyxl. ({e})"
            )
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        headers = ["" if h is None else str(h) for h in rows[0]]
        for vals in rows[1:]:
            if vals is None:
                continue
            row = {headers[i]: vals[i] if i < len(vals) else None for i in range(len(headers))}
            if all((v is None or str(v).strip() == "") for v in row.values()):
                continue
            yield row
        return

    raise SystemExit(f"ERROR: Unsupported input file type: {path}")


def _login_and_csrf(client: ApiClient, username: str, password: str) -> str:
    login = client.request_json(
        "POST",
        "/api/auth/login",
        body={"username": username, "password": password},
    )
    if not login.get("ok"):
        raise SystemExit(f"ERROR: login failed: {login}")
    csrf = ((login.get("identity") or {}).get("csrf")) if isinstance(login.get("identity"), dict) else None
    if not csrf:
        csrf_resp = client.request_json("GET", "/api/auth/csrf")
        csrf = csrf_resp.get("csrf_token")
    if not csrf:
        raise SystemExit("ERROR: could not obtain CSRF token after login")
    return str(csrf)


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(
        description="Create customers for each unique customer name in a print-plate database file (CSV/TSV/XLSX)."
    )
    p.add_argument("input_path", help="Path to plate database file")
    p.add_argument(
        "--base-url",
        default=os.getenv("API_BASE_URL", "http://localhost:8000"),
        help="API base URL (default: http://localhost:8000 or API_BASE_URL)",
    )
    p.add_argument("--username", default=os.getenv("API_USERNAME"), help="Login username (or API_USERNAME)")
    p.add_argument("--password", default=os.getenv("API_PASSWORD"), help="Login password (or API_PASSWORD)")
    p.add_argument(
        "--delimiter",
        default=None,
        help="CSV delimiter (optional). If omitted, attempts to sniff; use '\\t' for TSV.",
    )
    p.add_argument(
        "--out",
        default=None,
        help="Optional JSON output path for mapping {customer_name: customer_id}",
    )

    args = p.parse_args(argv)

    username = args.username
    if not username:
        raise SystemExit("ERROR: --username (or API_USERNAME) is required")

    password = args.password
    if not password:
        password = getpass.getpass("Password: ")
    if not password:
        raise SystemExit("ERROR: empty password not allowed")

    delimiter = args.delimiter
    if delimiter == "\\t":
        delimiter = "\t"

    rows = list(_iter_rows(args.input_path, delimiter))
    if not rows:
        raise SystemExit("ERROR: no rows found in input")

    names: list[str] = []
    seen = set[str]()
    for r in rows:
        name = _get(r, ["customer", "customer_name", "customername", "client", "account", "company", "customer name"])
        name = (name or "").strip()
        if not name:
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        names.append(name)

    if not names:
        raise SystemExit("ERROR: could not find any customer names in input (expected a Customer/Customer Name column)")

    client = ApiClient(args.base_url)
    csrf = _login_and_csrf(client, username, password)

    # Load existing customers once (for skipping + for used code uniqueness).
    existing_items = (client.request_json("GET", "/api/customers").get("items") or [])  # type: ignore[union-attr]
    existing_by_name = {str(c.get("name", "")).strip().lower(): c for c in existing_items if isinstance(c, dict)}
    used_codes = {str(c.get("code", "")).strip().upper() for c in existing_items if isinstance(c, dict) and c.get("code")}

    mapping: dict[str, str] = {}
    created = 0
    skipped = 0

    for name in names:
        existing = existing_by_name.get(name.strip().lower())
        if existing and existing.get("id"):
            mapping[name] = str(existing["id"])
            skipped += 1
            continue

        code = _derive_customer_code(name, used_codes)
        payload = _default_customer_payload(name, code)
        resp = client.request_json(
            "POST",
            "/api/customers",
            body=payload,
            headers={"x-csrf-token": csrf},
        )
        cust = resp.get("customer") if isinstance(resp, dict) else None
        cid = cust.get("id") if isinstance(cust, dict) else None
        if not cid:
            raise SystemExit(f"ERROR: unexpected response creating customer {name!r}: {json.dumps(resp)}")
        mapping[name] = str(cid)
        created += 1

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            json.dump(mapping, f, indent=2, sort_keys=True)

    print(
        json.dumps(
            {
                "ok": True,
                "unique_names": len(names),
                "created": created,
                "skipped_existing": skipped,
                "mapping": mapping if not args.out else f"(written to {args.out})",
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

