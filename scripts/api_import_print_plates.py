from __future__ import annotations

import argparse
import csv
import getpass
import json
import os
import re
import sys
from dataclasses import dataclass
from http.cookiejar import CookieJar
from typing import Any, Dict, Iterable, Optional, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urljoin
from urllib.request import HTTPCookieProcessor, Request, build_opener


@dataclass
class ApiClient:
    base_url: str

    def __post_init__(self) -> None:
        if not self.base_url.endswith("/"):
            self.base_url += "/"
        self.cookies = CookieJar()
        self.opener = build_opener(HTTPCookieProcessor(self.cookies))

    def request_json(
        self,
        method: str,
        path: str,
        *,
        body: Optional[Dict[str, Any]] = None,
        headers: Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        url = urljoin(self.base_url, path.lstrip("/"))
        data = None
        req_headers = {"accept": "application/json"}
        if headers:
            req_headers.update(headers)
        if body is not None:
            data = json.dumps(body).encode("utf-8")
            req_headers.setdefault("content-type", "application/json")
        req = Request(url=url, data=data, method=method.upper(), headers=req_headers)
        try:
            with self.opener.open(req, timeout=60) as resp:
                raw = resp.read().decode("utf-8")
                if not raw:
                    return {}
                return json.loads(raw)
        except HTTPError as e:
            raw = ""
            try:
                raw = e.read().decode("utf-8")
            except Exception:
                pass
            raise RuntimeError(f"HTTP {e.code} {e.reason} for {method} {url}: {raw}") from e
        except URLError as e:
            raise RuntimeError(f"Network error for {method} {url}: {e}") from e


def _norm_key(k: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (k or "").strip().lower())


def _get(row: dict[str, Any], keys: Iterable[str]) -> str:
    want = {_norm_key(k) for k in keys}
    for rk, rv in row.items():
        if _norm_key(rk) in want:
            return "" if rv is None else str(rv).strip()
    return ""


def _iter_rows(path: str, delimiter: str | None) -> Iterable[dict[str, Any]]:
    ext = os.path.splitext(path)[1].lower()
    if ext in [".csv", ".tsv", ".txt"]:
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            if delimiter is None:
                sample = f.read(4096)
                f.seek(0)
                sniff = csv.Sniffer()
                try:
                    dialect = sniff.sniff(sample, delimiters=[",", "\t", ";", "|"])
                except Exception:
                    dialect = csv.excel
                delim = getattr(dialect, "delimiter", ",")
            else:
                delim = delimiter
            reader = csv.DictReader(f, delimiter=delim)
            for r in reader:
                if not r:
                    continue
                yield r
        return

    if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        try:
            import openpyxl  # type: ignore
        except Exception as e:
            raise SystemExit(
                f"ERROR: {path} looks like an Excel file, but openpyxl isn't available. "
                f"Either export to CSV/TSV, or install openpyxl. ({e})"
            )
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        headers = ["" if h is None else str(h) for h in rows[0]]
        for vals in rows[1:]:
            if vals is None:
                continue
            row = {headers[i]: vals[i] if i < len(vals) else None for i in range(len(headers))}
            if all((v is None or str(v).strip() == "") for v in row.values()):
                continue
            yield row
        return

    raise SystemExit(f"ERROR: Unsupported input file type: {path}")


def _login_and_csrf(client: ApiClient, username: str, password: str) -> str:
    login = client.request_json(
        "POST",
        "/api/auth/login",
        body={"username": username, "password": password},
    )
    if not login.get("ok"):
        raise SystemExit(f"ERROR: login failed: {login}")
    csrf = ((login.get("identity") or {}).get("csrf")) if isinstance(login.get("identity"), dict) else None
    if not csrf:
        csrf_resp = client.request_json("GET", "/api/auth/csrf")
        csrf = csrf_resp.get("csrf_token")
    if not csrf:
        raise SystemExit("ERROR: could not obtain CSRF token after login")
    return str(csrf)


def _load_customer_map(client: ApiClient) -> dict[str, str]:
    resp = client.request_json("GET", "/api/customers")
    items = resp.get("items")
    if not isinstance(items, list):
        return {}
    out: dict[str, str] = {}
    for c in items:
        if not isinstance(c, dict):
            continue
        name = str(c.get("name", "")).strip()
        cid = c.get("id")
        if name and cid:
            out[name.lower()] = str(cid)
    return out


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(
        description="Upsert print plates from a plate database file into /api/admin/rate-cards/plates/{customer_id}/{plate_code}."
    )
    p.add_argument("input_path", help="Path to plate database file")
    p.add_argument(
        "--base-url",
        default=os.getenv("API_BASE_URL", "http://localhost:8000"),
        help="API base URL (default: http://localhost:8000 or API_BASE_URL)",
    )
    p.add_argument("--username", default=os.getenv("API_USERNAME"), help="Login username (or API_USERNAME)")
    p.add_argument("--password", default=os.getenv("API_PASSWORD"), help="Login password (or API_PASSWORD)")
    p.add_argument(
        "--delimiter",
        default=None,
        help="CSV delimiter (optional). If omitted, attempts to sniff; use '\\t' for TSV.",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and summarize without calling the API.",
    )

    args = p.parse_args(argv)

    username = args.username
    if not username:
        raise SystemExit("ERROR: --username (or API_USERNAME) is required")

    password = args.password
    if not password:
        password = getpass.getpass("Password: ")
    if not password:
        raise SystemExit("ERROR: empty password not allowed")

    delimiter = args.delimiter
    if delimiter == "\\t":
        delimiter = "\t"

    rows = list(_iter_rows(args.input_path, delimiter))
    if not rows:
        raise SystemExit("ERROR: no rows found in input")

    # (customer_name_lower, plate_code) -> payload
    plates: dict[Tuple[str, str], dict[str, Any]] = {}
    skipped_rows = 0
    for r in rows:
        customer_name = _get(r, ["customer", "customer_name", "customername", "client", "account", "company", "customer name"]).strip()
        plate_code = _get(r, ["plate_code", "platecode", "plate", "plate id", "plate_id", "plateid", "code"]).strip()
        if not customer_name or not plate_code:
            skipped_rows += 1
            continue

        description = _get(r, ["description", "desc", "plate_description", "name"]).strip() or None
        cylinder = _get(r, ["cylinder", "cyl", "repeat", "circumference"]).strip() or None

        plates[(customer_name.lower(), plate_code)] = {
            "description": description,
            "cylinder": cylinder,
        }

    if not plates:
        raise SystemExit("ERROR: no usable plate rows found (need customer name + plate code columns)")

    client = ApiClient(args.base_url)
    csrf = _login_and_csrf(client, username, password)

    cust_map = _load_customer_map(client)
    missing_customers = sorted({cn for (cn, _pc) in plates.keys() if cn not in cust_map})
    if missing_customers:
        raise SystemExit(
            "ERROR: some customers referenced by the plate file do not exist yet. "
            "Ensure those customer names exist in the app (MYOB customer sync, UI, or POST /api/customers). "
            "Missing: "
            + ", ".join(missing_customers[:20])
            + (" ..." if len(missing_customers) > 20 else "")
        )

    upserted = 0
    if not args.dry_run:
        for (cust_name_l, plate_code), payload in plates.items():
            customer_id = cust_map[cust_name_l]
            path = f"/api/admin/rate-cards/plates/{quote(customer_id)}/{quote(plate_code)}"
            _ = client.request_json("PUT", path, body=payload, headers={"x-csrf-token": csrf})
            upserted += 1

    print(
        json.dumps(
            {
                "ok": True,
                "distinct_plate_keys": len(plates),
                "skipped_rows_missing_required_fields": skipped_rows,
                "upserted": upserted,
                "dry_run": bool(args.dry_run),
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

