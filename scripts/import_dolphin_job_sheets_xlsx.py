#!/usr/bin/env python3
"""Extract Dolphin job-sheet XLSX data and match to existing order line items.

Default mode is non-destructive:
- parse all workbooks
- score candidate matches against existing Dolphin-imported order lines
- write review outputs (JSON + CSV)

Use ``--apply-link-missing-job-sheets`` for an optional DB write pass that only:
- links a fresh import-draft job sheet to matched lines that currently have no job sheet
- leaves existing linked job sheets untouched
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
import uuid
import warnings
from dataclasses import asdict, dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from sqlalchemy import select

# Repo root: scripts/ -> parent
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.db.models.domain import Customer, Order, OrderItem, Product, ProductVersion  # noqa: E402
from app.db.myob_import_placeholders import MYOB_DRAFT_PLACEHOLDER_PRODUCT_ID  # noqa: E402
from app.db.models.domain import JobSheet  # noqa: E402
from app.db.session import SessionLocal  # noqa: E402
from app.exceptions import DomainError  # noqa: E402
from app.job_sheets import service as job_sheets_service  # noqa: E402
from app.products.schemas import SpecPayload  # noqa: E402
from app.products.service import create_product_v1_in_session  # noqa: E402

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    module="openpyxl.worksheet._reader",
)

try:
    import openpyxl  # type: ignore
except Exception as e:  # pragma: no cover - explicit startup guidance
    raise SystemExit(
        "openpyxl is required. Install in the project venv:\n"
        "  .venv/bin/pip install openpyxl\n"
        f"Import error: {e}"
    )


LABEL_ALIASES: dict[str, tuple[str, ...]] = {
    "item": ("item",),
    "item_code": ("item code", "itemcode"),
    "customer": ("customer",),
    "batch_no": ("batch no", "batch number", "batch"),
    "due_date": ("due date", "due"),
    "created": ("created", "created date"),
    "total_rolls": ("total rolls",),
    "total_kgs": ("total kgs", "total kg"),
    "total_mts": ("total mts", "total mts.", "total meters", "total metres"),
    "print_type": ("print type", "printing type"),
    "roll_type": ("roll type",),
    "slit": ("slit", "slits"),
}


def norm_text(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip()).strip()


def fold(v: Any) -> str:
    s = norm_text(v).lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def parse_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return f if math.isfinite(f) else None
    s = str(v).replace(",", "").strip()
    if not s:
        return None
    try:
        f = float(s)
        return f if math.isfinite(f) else None
    except ValueError:
        return None


def parse_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = norm_text(v)
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d %b %Y",
        "%d %B %Y",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


@dataclass
class ExtractedSheet:
    file_name: str
    file_stem: str
    sheet_name: str
    item: str | None
    item_code: str | None
    customer: str | None
    batch_no: str | None
    due_date: str | None
    created_date: str | None
    total_rolls: float | None
    total_kgs: float | None
    total_mts: float | None
    print_type: str | None
    roll_type: str | None
    slit: str | None
    notes: str | None


@dataclass
class CandidateLine:
    order_item_id: str
    order_id: str
    order_code: str | None
    order_date: str | None
    customer_id: str
    customer_name: str | None
    line_index: int
    line_kind: str
    myob_item_number: str | None
    import_line_description: str | None
    import_ship_quantity: float | None
    import_quantity_unit: str | None
    import_qty_type: str | None
    import_requires_job_sheet: bool
    job_sheet_id: str | None


@dataclass
class ScoredCandidate:
    candidate: CandidateLine
    score: float
    reasons: list[str]


def _iter_top_cells(ws, *, max_row: int = 40, max_col: int = 20):
    for r in range(1, min(max_row, ws.max_row) + 1):
        for c in range(1, min(max_col, ws.max_column) + 1):
            yield r, c, ws.cell(r, c).value


def _extract_label_values(ws) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, aliases in LABEL_ALIASES.items():
        found = False
        for r, c, raw in _iter_top_cells(ws):
            text = fold(raw)
            if not text:
                continue
            if text in aliases:
                right = ws.cell(r, c + 1).value
                if right is not None and norm_text(right):
                    out[key] = right
                    found = True
                    break
                # fallback: search rightward for first non-empty value
                for c2 in range(c + 1, min(c + 8, ws.max_column) + 1):
                    v = ws.cell(r, c2).value
                    if v is not None and norm_text(v):
                        out[key] = v
                        found = True
                        break
                if found:
                    break
        if not found:
            out[key] = None
    return out


def _extract_notes(ws) -> str | None:
    best = ""
    for r in range(1, min(30, ws.max_row) + 1):
        vals: list[str] = []
        for c in range(1, min(8, ws.max_column) + 1):
            v = ws.cell(r, c).value
            t = norm_text(v)
            if t:
                vals.append(t)
        if not vals:
            continue
        row_text = " ".join(vals)
        f = fold(row_text)
        if any(
            k in f
            for k in (
                "item specifications",
                "printing",
                "resin mix",
                "components",
                "extrusion specs",
                "labels",
            )
        ):
            continue
        # Generic machine-setup heading that repeats across sheets and does not map to spec fields.
        if (
            "extruder start time" in f
            and "screw speed" in f
            and "nip speed" in f
            and "blower" in f
            and "nip roller tension speed" in f
        ):
            continue
        if len(row_text) > len(best) and len(row_text) >= 24:
            best = row_text
    return best or None


def extract_sheet(path: Path) -> ExtractedSheet:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    kv = _extract_label_values(ws)
    notes = _extract_notes(ws)
    return ExtractedSheet(
        file_name=path.name,
        file_stem=path.stem,
        sheet_name=ws.title,
        item=norm_text(kv.get("item")) or None,
        item_code=norm_text(kv.get("item_code")) or None,
        customer=norm_text(kv.get("customer")) or None,
        batch_no=norm_text(kv.get("batch_no")) or None,
        due_date=parse_date(kv.get("due_date")).isoformat() if parse_date(kv.get("due_date")) else None,
        created_date=parse_date(kv.get("created")).isoformat() if parse_date(kv.get("created")) else None,
        total_rolls=parse_float(kv.get("total_rolls")),
        total_kgs=parse_float(kv.get("total_kgs")),
        total_mts=parse_float(kv.get("total_mts")),
        print_type=norm_text(kv.get("print_type")) or None,
        roll_type=norm_text(kv.get("roll_type")) or None,
        slit=norm_text(kv.get("slit")) or None,
        notes=notes,
    )


def load_candidate_lines(db) -> list[CandidateLine]:
    stmt = (
        select(OrderItem, Order, Customer)
        .join(Order, Order.id == OrderItem.order_id)
        .join(Customer, Customer.id == Order.customer_id)
        .where(OrderItem.line_kind == "myob_import")
        .where(Order.import_source == "DOLPHIN_TSV")
    )
    out: list[CandidateLine] = []
    for oi, o, c in db.execute(stmt).all():
        out.append(
            CandidateLine(
                order_item_id=str(oi.id),
                order_id=str(o.id),
                order_code=str(getattr(o, "code", "") or "") or None,
                order_date=str(getattr(o, "order_date", "") or "") or None,
                customer_id=str(c.id),
                customer_name=str(getattr(c, "name", "") or "") or None,
                line_index=int(getattr(oi, "line_index", 0) or 0),
                line_kind=str(getattr(oi, "line_kind", "") or ""),
                myob_item_number=str(getattr(oi, "myob_item_number", "") or "") or None,
                import_line_description=str(getattr(oi, "import_line_description", "") or "") or None,
                import_ship_quantity=float(getattr(oi, "import_ship_quantity", 0.0))
                if getattr(oi, "import_ship_quantity", None) is not None
                else None,
                import_quantity_unit=str(getattr(oi, "import_quantity_unit", "") or "") or None,
                import_qty_type=str(getattr(oi, "import_qty_type", "") or "") or None,
                import_requires_job_sheet=bool(getattr(oi, "import_requires_job_sheet", False)),
                job_sheet_id=str(getattr(oi, "job_sheet_id", "") or "") or None,
            )
        )
    return out


def _token_overlap(a: str, b: str) -> float:
    ta = {t for t in fold(a).split() if len(t) >= 2}
    tb = {t for t in fold(b).split() if len(t) >= 2}
    if not ta or not tb:
        return 0.0
    inter = len(ta & tb)
    denom = max(len(ta), len(tb))
    return inter / denom if denom else 0.0


def _clean_item_value(v: str | None) -> str:
    s = norm_text(v)
    if not s:
        return ""
    sf = fold(s)
    if any(x in sf for x in ("#value", "#ref", "#name", "error", "na")):
        return ""
    return s


def _description_tokens(desc: str | None) -> list[str]:
    s = fold(desc)
    if not s:
        return []
    return [t for t in re.split(r"[^a-z0-9]+", s) if t]


def score_candidate(sheet: ExtractedSheet, cand: CandidateLine) -> ScoredCandidate:
    score = 0.0
    reasons: list[str] = []

    sheet_customer = fold(sheet.customer)
    cand_customer = fold(cand.customer_name)
    if sheet_customer and cand_customer:
        if sheet_customer == cand_customer:
            score += 35
            reasons.append("customer_exact")
        elif sheet_customer in cand_customer or cand_customer in sheet_customer:
            score += 20
            reasons.append("customer_partial")

    desc_fold = fold(cand.import_line_description)
    desc_tokens = _description_tokens(cand.import_line_description)

    # Primary: workbook Item should match the START of order line description.
    item = _clean_item_value(sheet.item)
    if item and desc_fold:
        item_fold = fold(item)
        if desc_fold.startswith(item_fold):
            score += 130
            reasons.append("item_matches_description_prefix")
        elif item_fold in desc_fold:
            score += 45
            reasons.append("item_found_in_description")

    # Fallback: when Item formula is broken/missing, Item Code can match the second segment/token.
    item_code = _clean_item_value(sheet.item_code)
    if item_code:
        item_code_fold = fold(item_code)
        if len(desc_tokens) >= 2 and desc_tokens[1] == item_code_fold:
            score += 110
            reasons.append("item_code_matches_description_second_token")
        elif item_code_fold in desc_tokens:
            score += 60
            reasons.append("item_code_matches_description_token")
        elif desc_fold and item_code_fold in desc_fold:
            score += 30
            reasons.append("item_code_found_in_description")

    # Keep myob_item_number as a secondary hint.
    if cand.myob_item_number:
        num_fold = fold(cand.myob_item_number)
        if item and num_fold == fold(item):
            score += 30
            reasons.append("item_equals_myob_item_number")
        if item_code and num_fold == fold(item_code):
            score += 25
            reasons.append("item_code_equals_myob_item_number")

    if cand.myob_item_number and fold(cand.myob_item_number) in fold(sheet.file_stem):
        score += 20
        reasons.append("filename_contains_item")

    if sheet.total_rolls is not None and cand.import_ship_quantity is not None:
        q0 = float(sheet.total_rolls)
        q1 = float(cand.import_ship_quantity)
        if q0 > 0 and q1 > 0:
            rel = abs(q0 - q1) / max(q0, q1)
            if rel <= 0.01:
                score += 20
                reasons.append("qty_rolls_close_1pct")
            elif rel <= 0.05:
                score += 10
                reasons.append("qty_rolls_close_5pct")

    text_sheet = " ".join(x for x in [sheet.notes, sheet.file_stem, sheet.item, sheet.item_code] if x)
    text_cand = " ".join(x for x in [cand.import_line_description, cand.myob_item_number] if x)
    ov = _token_overlap(text_sheet, text_cand)
    if ov >= 0.5:
        score += 25
        reasons.append("desc_overlap_high")
    elif ov >= 0.25:
        score += 12
        reasons.append("desc_overlap_medium")

    if cand.import_requires_job_sheet:
        score += 8
        reasons.append("line_requires_job_sheet")

    return ScoredCandidate(candidate=cand, score=score, reasons=reasons)


def choose_match(
    scored: list[ScoredCandidate], *, high_threshold: float, ambiguous_gap: float
) -> tuple[str, ScoredCandidate | None, list[ScoredCandidate]]:
    if not scored:
        return ("unmatched", None, [])
    ranked = sorted(scored, key=lambda x: x.score, reverse=True)
    top = ranked[0]
    second = ranked[1] if len(ranked) > 1 else None
    if top.score < high_threshold:
        return ("unmatched", None, ranked[:5])
    if second is not None and (top.score - second.score) < ambiguous_gap:
        return ("ambiguous", top, ranked[:5])
    return ("matched", top, ranked[:5])


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    keys = sorted({k for r in rows for k in r.keys()})
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def _progress(label: str, done: int, total: int) -> None:
    total_safe = max(total, 1)
    pct = (float(done) / float(total_safe)) * 100.0
    print(f"[{label}] {done}/{total} ({pct:5.1f}%)")


def _truthy(v: Any) -> bool:
    s = str(v or "").strip().lower()
    return s in ("1", "true", "t", "yes", "y")


def _parse_item_dimensions(item: str | None) -> tuple[int | None, int | None, int | None]:
    s = norm_text(item)
    if not s:
        return (None, None, None)
    sf = fold(s)
    # Centerfold convention like: CF1000(2000)200050
    # - first number = layflat cue (1000)
    # - parenthesized number = actual centerfold width (2000)
    # We import base width as centerfold width so layflat resolves to half-width (1000).
    m_cf = re.search(r"cf\s*(\d{2,4})\s*\(\s*(\d{2,5})\s*\)\s*(\d{2,6})", sf, flags=re.IGNORECASE)
    if m_cf:
        centerfold_width = int(m_cf.group(2))
        tail = m_cf.group(3)
        if len(tail) >= 4:
            length = int(tail[:-2])
            gauge = int(tail[-2:])
        else:
            length = int(tail)
            gauge = None
        return (centerfold_width, length, gauge)
    nums = [int(x) for x in re.findall(r"\d{2,4}", s)]
    if len(nums) >= 3:
        return (nums[0], nums[1], nums[2])
    if len(nums) == 2:
        return (nums[0], nums[1], None)
    if len(nums) == 1:
        return (nums[0], None, None)
    return (None, None, None)


def _spec_from_row_for_promotion(row: dict[str, Any], *, fallback_qty_unit: str = "rolls") -> SpecPayload:
    item = norm_text(row.get("item"))
    print_type = fold(row.get("print_type"))
    width, length, gauge = _parse_item_dimensions(item or row.get("file_stem"))
    prefix = fold(item)[:3]

    if prefix.startswith("t"):
        product_type = "Tube"
        length_units = "Continuous"
        base_length = None
    elif prefix.startswith("s"):
        product_type = "Sheet"
        length_units = "mm"
        base_length = length if length and length > 0 else 1000
    elif prefix.startswith("cf"):
        product_type = "Centerfold"
        length_units = "mm"
        base_length = length if length and length > 0 else 1000
    else:
        product_type = "Bag"
        length_units = "mm"
        base_length = length if length and length > 0 else 1000

    finish_mode = "Rolls"
    if str(fallback_qty_unit or "").strip().lower() in ("cartons", "bags"):
        finish_mode = "Cartons"
        if base_length is None:
            base_length = 1000

    slit_raw = fold(row.get("slit"))
    if "both" in slit_raw or "2 side" in slit_raw or "two side" in slit_raw:
        slit_val = "both_sides"
    elif "one" in slit_raw or "single" in slit_raw:
        slit_val = "one_side"
    elif "middle" in slit_raw or "center" in slit_raw or "centre" in slit_raw:
        slit_val = "middle"
    else:
        slit_val = "none"

    spec_dict: dict[str, Any] = {
        "identity": {
            "product_type": product_type,
            "finish_mode": finish_mode,
            "industry_flags": [],
            "notes": None,
            "customer_code": norm_text(row.get("item_code")) or None,
        },
        "dimensions": {
            "base_width_mm": width if width and width > 0 else 300,
            "width_tolerance_mm": None,
            "base_length_mm": base_length,
            "thickness_um": gauge if gauge and gauge > 0 else 50,
            "geometry": "Flat",
            "gusset_mm": None,
            "ufilm_left_width_mm": None,
            "ufilm_right_width_mm": None,
            "length_units": length_units,
        },
        "formulation": {
            "blend_type": "Custom",
            "blend": [{"resin_code": "LD", "pct": 100}],
            "colour": None,
            "colour_components": [],
            "additives": [],
        },
        "printing": {
            "method": "None" if (not print_type or print_type == "none") else "Inline",
            "num_colours": 0,
            "print_description": None,
            "ink_codes": [],
            "plate_codes": [],
            "side": None,
            "artwork_refs": [],
            "artwork_files": [],
            "front_ink_plate": [],
            "back_ink_plate": [],
            "cylinder_size_mm": None,
            "barcode": None,
            "print_position_notes": None,
            "plates_around": None,
            "plates_across": None,
            "seal_type": None,
            "eye_spot": None,
        },
        "quality_expectations": {"flags": [], "known_issues": None},
        "run_requirements": {
            "preferred_extruders": [],
            "preferred_printer": None,
            "preferred_converter": None,
            # Backend schema does not store explicit 1-up; "none" is effective 1-up/default.
            "run_up": "none",
            "slit": slit_val,
            "treat_inside_outside": "none",
            "inline_perforation": False,
            "hole_punched": False,
            "inline_seal": False,
            "notes": None,
        },
        "packaging": {
            "pack_mode": finish_mode,
            "core_type": "None",
            "core_policy": "Include",
            "bags_per_carton": 20 if finish_mode == "Cartons" else None,
            "pallet_type": "None",
            "notes": None,
        },
        "tool_requirements": [],
        "sensor_qc_config": None,
        "wi_mappings": None,
    }
    return SpecPayload(**spec_dict)


def _promote_placeholder_job_sheet(
    db,
    *,
    js: JobSheet,
    row: dict[str, Any],
    created_by: str,
    fallback_qty_unit: str,
) -> None:
    spec = _spec_from_row_for_promotion(row, fallback_qty_unit=fallback_qty_unit)
    try:
        new_product, version = create_product_v1_in_session(
            db,
            customer_id=str(js.customer_id),
            spec=spec,
            created_by=created_by,
        )
    except DomainError as e:
        # When many rows share effectively identical spec-derived codes, fall back to a generated
        # unique internal product code. Customer-facing code remains on spec.identity.customer_code.
        if "unique customer-facing product code" not in str(e):
            raise
        spec_dict = spec.model_dump() if hasattr(spec, "model_dump") else spec.dict()
        fallback_desc = norm_text(row.get("item")) or norm_text(row.get("item_code")) or "Imported Dolphin job sheet"
        for _ in range(50):
            code = f"IMP-{uuid.uuid4().hex[:8].upper()}"[:32]
            exists = db.scalar(select(Product.id).where(Product.code == code))
            if exists:
                continue
            p = Product(code=code, description=fallback_desc[:255], customer_id=str(js.customer_id))
            db.add(p)
            db.flush()
            v = ProductVersion(
                product_id=str(p.id),
                version_number=1,
                created_by=created_by or "system",
                spec_payload=spec_dict,
            )
            db.add(v)
            db.flush()
            p.active_version_id = str(v.id)
            db.add(p)
            db.flush()
            new_product, version = p, v
            break
        else:  # pragma: no cover
            raise DomainError("Failed to allocate fallback internal product code for Dolphin import")
    js.product_id = str(new_product.id)
    js.product_version_id = str(version.id)
    js.is_import_draft = False
    cf_desc = norm_text(row.get("item")) or norm_text(row.get("item_code"))
    if cf_desc:
        js.customer_facing_description = cf_desc
    db.add(js)
    db.flush()
    job_sheets_service.finalize_import_draft_job_sheet_after_spec_save(db, str(js.id))


def _apply_from_matched_csv(
    db,
    *,
    matched_csv_path: Path,
    created_by: str,
    dry_run: bool = False,
) -> dict[str, int]:
    if not matched_csv_path.is_file():
        raise FileNotFoundError(f"matched csv not found: {matched_csv_path}")
    with matched_csv_path.open("r", newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    created = 0
    promoted_placeholder = 0
    skipped_has_job_sheet = 0
    skipped_no_requires = 0
    missing_order_item = 0
    bad_rows = 0

    step = max(1, len(rows) // 20) if rows else 1
    for i, row in enumerate(rows, start=1):
        oid = str(row.get("order_item_id") or "").strip()
        if not oid:
            bad_rows += 1
            continue
        oi = db.get(OrderItem, oid)
        if oi is None:
            missing_order_item += 1
            continue
        existing_js: JobSheet | None = None
        if getattr(oi, "job_sheet_id", None):
            existing_js = db.get(JobSheet, str(getattr(oi, "job_sheet_id")))
            is_placeholder = bool(
                existing_js
                and (
                    bool(getattr(existing_js, "is_import_draft", False))
                    or str(getattr(existing_js, "product_id", "")) == str(MYOB_DRAFT_PLACEHOLDER_PRODUCT_ID)
                )
            )
            if is_placeholder:
                if not dry_run and existing_js is not None:
                    _promote_placeholder_job_sheet(
                        db,
                        js=existing_js,
                        row=row,
                        created_by=created_by,
                        fallback_qty_unit=str(getattr(oi, "import_quantity_unit", "rolls") or "rolls"),
                    )
                promoted_placeholder += 1
                if i == 1 or i == len(rows) or (i % step == 0):
                    _progress("apply", i, len(rows))
                continue
            skipped_has_job_sheet += 1
            continue
        # Prefer current DB flag; fallback to CSV flag when absent.
        requires = getattr(oi, "import_requires_job_sheet", None)
        if requires is None:
            requires = _truthy(row.get("import_requires_job_sheet"))
        if not bool(requires):
            skipped_no_requires += 1
            continue
        order = db.get(Order, str(getattr(oi, "order_id", "")))
        if order is None:
            missing_order_item += 1
            continue
        if not dry_run:
            js = job_sheets_service.create_myob_import_draft_job_sheet(
                db=db,
                customer_id=str(order.customer_id),
                quantity_value=float(getattr(oi, "import_ship_quantity", 0.0) or 0.0),
                quantity_unit=str(getattr(oi, "import_quantity_unit", "ea") or "ea"),
                qty_type=str(getattr(oi, "import_qty_type", "units") or "units"),
                unit_rate=float(getattr(oi, "import_unit_price", 0.0))
                if getattr(oi, "import_unit_price", None) is not None
                else None,
                line_total=float(getattr(oi, "import_line_total", 0.0))
                if getattr(oi, "import_line_total", None) is not None
                else None,
                created_by=created_by,
            )
            _promote_placeholder_job_sheet(
                db,
                js=js,
                row=row,
                created_by=created_by,
                fallback_qty_unit=str(getattr(oi, "import_quantity_unit", "rolls") or "rolls"),
            )
            oi.job_sheet_id = str(js.id)
            db.add(oi)
        created += 1
        if i == 1 or i == len(rows) or (i % step == 0):
            _progress("apply", i, len(rows))

    return {
        "rows_in_csv": len(rows),
        "created_job_sheets": created,
        "promoted_placeholder_job_sheets": promoted_placeholder,
        "skipped_has_job_sheet": skipped_has_job_sheet,
        "skipped_no_requires_job_sheet": skipped_no_requires,
        "missing_order_item": missing_order_item,
        "bad_rows": bad_rows,
    }


def _apply_link_missing_job_sheets(db, accepted: list[tuple[ExtractedSheet, ScoredCandidate]]) -> dict[str, int]:
    created = 0
    skipped_has_job_sheet = 0
    skipped_no_requires = 0
    missing_line = 0
    for _sheet, sc in accepted:
        cand = sc.candidate
        oi = db.get(OrderItem, cand.order_item_id)
        if oi is None:
            missing_line += 1
            continue
        if getattr(oi, "job_sheet_id", None):
            skipped_has_job_sheet += 1
            continue
        if not bool(getattr(oi, "import_requires_job_sheet", False)):
            skipped_no_requires += 1
            continue
        order = db.get(Order, str(getattr(oi, "order_id", "")))
        if order is None:
            missing_line += 1
            continue
        js = job_sheets_service.create_myob_import_draft_job_sheet(
            db=db,
            customer_id=str(order.customer_id),
            quantity_value=float(getattr(oi, "import_ship_quantity", 0.0) or 0.0),
            quantity_unit=str(getattr(oi, "import_quantity_unit", "ea") or "ea"),
            qty_type=str(getattr(oi, "import_qty_type", "units") or "units"),
            unit_rate=float(getattr(oi, "import_unit_price", 0.0))
            if getattr(oi, "import_unit_price", None) is not None
            else None,
            line_total=float(getattr(oi, "import_line_total", 0.0))
            if getattr(oi, "import_line_total", None) is not None
            else None,
            created_by="Dolphin XLSX import",
        )
        oi.job_sheet_id = str(js.id)
        db.add(oi)
        created += 1
    return {
        "created_job_sheets": created,
        "skipped_has_job_sheet": skipped_has_job_sheet,
        "skipped_no_requires_job_sheet": skipped_no_requires,
        "missing_order_item": missing_line,
    }


def main() -> int:
    p = argparse.ArgumentParser(
        description="Parse Dolphin XLSX job sheets and match to existing Dolphin-imported order lines."
    )
    p.add_argument(
        "--input-dir",
        default=str(ROOT / "scripts" / "dolphin-job-sheets"),
        help="Directory containing Dolphin XLSX files",
    )
    p.add_argument(
        "--output-dir",
        default=str(ROOT / "scripts" / "dolphin-job-sheets-output"),
        help="Directory for matcher outputs",
    )
    p.add_argument("--limit", type=int, default=0, help="Limit files processed (0 = all)")
    p.add_argument(
        "--high-threshold",
        type=float,
        default=90.0,
        help="Minimum score for auto-match",
    )
    p.add_argument(
        "--ambiguous-gap",
        type=float,
        default=15.0,
        help="Minimum top-vs-second score gap for auto-match",
    )
    p.add_argument(
        "--apply-link-missing-job-sheets",
        action="store_true",
        help="Write mode: create and link draft job sheets for matched lines with missing job_sheet_id.",
    )
    p.add_argument(
        "--apply-from-matched-csv",
        default="",
        help=(
            "Apply mode from an existing matched.csv path (skips XLSX re-parse/re-match). "
            "Useful to finalize a prior run."
        ),
    )
    p.add_argument(
        "--apply-dry-run",
        action="store_true",
        help="With --apply-from-matched-csv, calculate what would be created without DB writes.",
    )
    p.add_argument(
        "--created-by",
        default="Dolphin XLSX import",
        help="Audit value for created draft job sheets.",
    )
    args = p.parse_args()

    apply_csv_path = Path(str(args.apply_from_matched_csv).strip()) if str(args.apply_from_matched_csv).strip() else None
    if apply_csv_path is not None:
        stamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
        run_dir = Path(args.output_dir) / f"run-{stamp}"
        run_dir.mkdir(parents=True, exist_ok=True)
        with SessionLocal() as db:
            apply_summary = _apply_from_matched_csv(
                db,
                matched_csv_path=apply_csv_path,
                created_by=str(args.created_by or "Dolphin XLSX import"),
                dry_run=bool(args.apply_dry_run),
            )
            if not args.apply_dry_run:
                db.commit()
        summary = {
            "ok": True,
            "mode": "apply_from_matched_csv",
            "matched_csv_path": str(apply_csv_path),
            "output_dir": str(run_dir),
            "write_mode": not bool(args.apply_dry_run),
            "apply_summary": apply_summary,
        }
        (run_dir / "summary.json").write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")
        print(json.dumps(summary, indent=2, default=str))
        return 0

    in_dir = Path(args.input_dir)
    if not in_dir.is_dir():
        print(json.dumps({"ok": False, "error": f"input dir not found: {in_dir}"}))
        return 1
    files = sorted(in_dir.glob("*.xlsx"))
    if args.limit and args.limit > 0:
        files = files[: args.limit]
    if not files:
        print(json.dumps({"ok": False, "error": f"no .xlsx files found in: {in_dir}"}))
        return 1

    stamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    run_dir = Path(args.output_dir) / f"run-{stamp}"
    run_dir.mkdir(parents=True, exist_ok=True)

    extracted: list[ExtractedSheet] = []
    parse_errors: list[dict[str, str]] = []
    parse_step = max(1, len(files) // 20)
    for i, f in enumerate(files, start=1):
        try:
            extracted.append(extract_sheet(f))
        except Exception as e:  # noqa: BLE001
            parse_errors.append({"file_name": f.name, "error": str(e)})
        if i == 1 or i == len(files) or (i % parse_step == 0):
            _progress("parse", i, len(files))

    with SessionLocal() as db:
        candidates = load_candidate_lines(db)
        candidate_count = len(candidates)

        matched_rows: list[dict[str, Any]] = []
        ambiguous_rows: list[dict[str, Any]] = []
        unmatched_rows: list[dict[str, Any]] = []
        accepted_for_apply: list[tuple[ExtractedSheet, ScoredCandidate]] = []

        match_step = max(1, len(extracted) // 20)
        for i, sh in enumerate(extracted, start=1):
            pool = candidates
            if sh.customer:
                c_fold = fold(sh.customer)
                customer_filtered = [c for c in pool if fold(c.customer_name) == c_fold]
                if customer_filtered:
                    pool = customer_filtered
            scored = [score_candidate(sh, c) for c in pool]
            status, top, top5 = choose_match(
                scored, high_threshold=float(args.high_threshold), ambiguous_gap=float(args.ambiguous_gap)
            )
            base = asdict(sh)
            if status == "matched" and top is not None:
                row = {
                    **base,
                    "match_status": "matched",
                    "match_score": round(top.score, 3),
                    "match_reasons": ",".join(top.reasons),
                    "order_item_id": top.candidate.order_item_id,
                    "order_id": top.candidate.order_id,
                    "order_code": top.candidate.order_code,
                    "line_index": top.candidate.line_index,
                    "myob_item_number": top.candidate.myob_item_number,
                    "import_line_description": top.candidate.import_line_description,
                    "job_sheet_id": top.candidate.job_sheet_id,
                    "import_requires_job_sheet": top.candidate.import_requires_job_sheet,
                }
                matched_rows.append(row)
                accepted_for_apply.append((sh, top))
            elif status == "ambiguous":
                top_json = [
                    {
                        "score": round(sc.score, 3),
                        "reasons": sc.reasons,
                        "order_item_id": sc.candidate.order_item_id,
                        "order_code": sc.candidate.order_code,
                        "line_index": sc.candidate.line_index,
                        "myob_item_number": sc.candidate.myob_item_number,
                        "import_line_description": sc.candidate.import_line_description,
                        "customer_name": sc.candidate.customer_name,
                    }
                    for sc in top5
                ]
                ambiguous_rows.append(
                    {
                        **base,
                        "match_status": "ambiguous",
                        "top_candidates_json": json.dumps(top_json, ensure_ascii=False),
                    }
                )
            else:
                unmatched_rows.append(
                    {
                        **base,
                        "match_status": "unmatched",
                    }
                )
            if i == 1 or i == len(extracted) or (i % match_step == 0):
                _progress("match", i, len(extracted))

        apply_summary: dict[str, Any] = {}
        if args.apply_link_missing_job_sheets:
            apply_summary = _apply_link_missing_job_sheets(db, accepted_for_apply)
            db.commit()

    _write_csv(run_dir / "matched.csv", matched_rows)
    _write_csv(run_dir / "ambiguous.csv", ambiguous_rows)
    _write_csv(run_dir / "unmatched.csv", unmatched_rows)
    _write_csv(run_dir / "parse_errors.csv", parse_errors)

    summary = {
        "ok": True,
        "input_dir": str(in_dir),
        "output_dir": str(run_dir),
        "files_seen": len(files),
        "files_parsed": len(extracted),
        "parse_errors": len(parse_errors),
        "candidate_order_lines": candidate_count,
        "matched": len(matched_rows),
        "ambiguous": len(ambiguous_rows),
        "unmatched": len(unmatched_rows),
        "high_threshold": float(args.high_threshold),
        "ambiguous_gap": float(args.ambiguous_gap),
        "write_mode": bool(args.apply_link_missing_job_sheets),
        "apply_summary": apply_summary,
    }

    (run_dir / "summary.json").write_text(json.dumps(summary, indent=2, default=str), encoding="utf-8")
    print(json.dumps(summary, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

