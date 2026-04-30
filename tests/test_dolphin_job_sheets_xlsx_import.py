from __future__ import annotations

import uuid
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker

import app.db.models.domain  # noqa: F401
from app.db.models import Base
from app.db.models.domain import Customer, JobSheet, Order, OrderItem, Product, ProductVersion
from app.db.models.enums import OrderStatus
from app.db.myob_import_placeholders import (
    MYOB_DRAFT_INTERNAL_CUSTOMER_ID,
    MYOB_DRAFT_PLACEHOLDER_PRODUCT_ID,
    MYOB_DRAFT_PLACEHOLDER_VERSION_ID,
    MYOB_DRAFT_SPEC_PAYLOAD,
)
from scripts.import_dolphin_job_sheets_xlsx import (
    CandidateLine,
    _apply_from_matched_csv,
    choose_match,
    extract_sheet,
    score_candidate,
)


def _seed_myob_draft_placeholders(db) -> None:
    if db.get(Customer, MYOB_DRAFT_INTERNAL_CUSTOMER_ID):
        return
    db.add(
        Customer(
            id=MYOB_DRAFT_INTERNAL_CUSTOMER_ID,
            name="Internal (MYOB import placeholder)",
            status="Active",
            contacts={},
            delivery_addresses={},
            delivery_preferences={},
        )
    )
    p = Product(
        id=MYOB_DRAFT_PLACEHOLDER_PRODUCT_ID,
        code="__MYOB_IMPORT__",
        description="Placeholder for MYOB import draft job sheets",
        customer_id=MYOB_DRAFT_INTERNAL_CUSTOMER_ID,
    )
    db.add(p)
    pv = ProductVersion(
        id=MYOB_DRAFT_PLACEHOLDER_VERSION_ID,
        product_id=MYOB_DRAFT_PLACEHOLDER_PRODUCT_ID,
        version_number=1,
        created_by="test",
        spec_payload=dict(MYOB_DRAFT_SPEC_PAYLOAD),
    )
    db.add(pv)
    p.active_version_id = MYOB_DRAFT_PLACEHOLDER_VERSION_ID
    db.add(p)
    db.commit()


def _write_sample_workbook(path: Path, *, printing_label: str = "Print Type") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1).value = "Item"
    ws.cell(1, 2).value = "PB30045075"
    ws.cell(1, 5).value = "Customer"
    ws.cell(1, 6).value = "BLACKWOODS"
    ws.cell(2, 1).value = "Item Code"
    ws.cell(2, 2).value = "LCW77521"
    ws.cell(2, 5).value = "Batch No"
    ws.cell(2, 6).value = "951578"
    ws.cell(2, 9).value = "Due Date"
    ws.cell(2, 10).value = "2026-03-18 00:00:00"
    ws.cell(3, 1).value = "Please ensure no creasing and strong seals."
    ws.cell(10, 1).value = "Total Rolls"
    ws.cell(10, 2).value = 5
    ws.cell(10, 4).value = "Total Kgs"
    ws.cell(10, 5).value = 101.25
    ws.cell(10, 7).value = "Total Mts"
    ws.cell(10, 8).value = 2430
    ws.cell(13, 1).value = printing_label
    ws.cell(13, 2).value = "None"
    wb.save(path)


def _candidate(
    *,
    item_number: str,
    customer_name: str = "BLACKWOODS",
    quantity: float = 5.0,
    requires_job_sheet: bool = True,
) -> CandidateLine:
    return CandidateLine(
        order_item_id="oi-1",
        order_id="o-1",
        order_code="INV-001",
        order_date="2026-03-01",
        customer_id="c-1",
        customer_name=customer_name,
        line_index=0,
        line_kind="myob_import",
        myob_item_number=item_number,
        import_line_description="PB30045075 heavy duty bag",
        import_ship_quantity=quantity,
        import_quantity_unit="rolls",
        import_qty_type="total_rolls",
        import_requires_job_sheet=requires_job_sheet,
        job_sheet_id=None,
    )


def test_extract_sheet_reads_primary_labels(tmp_path: Path) -> None:
    p = tmp_path / "PB30045075.xlsx"
    _write_sample_workbook(p, printing_label="Print Type")
    out = extract_sheet(p)
    assert out.item == "PB30045075"
    assert out.item_code == "LCW77521"
    assert out.customer == "BLACKWOODS"
    assert out.batch_no == "951578"
    assert out.due_date == "2026-03-18"
    assert out.total_rolls == 5
    assert out.total_kgs == 101.25
    assert out.total_mts == 2430
    assert out.print_type == "None"
    assert out.notes is not None


def test_extract_sheet_handles_printing_type_variant_label(tmp_path: Path) -> None:
    p = tmp_path / "T300200-20KG.xlsx"
    _write_sample_workbook(p, printing_label="Printing Type")
    out = extract_sheet(p)
    assert out.print_type == "None"


def test_scoring_and_choose_match_prefers_strong_item_number() -> None:
    class SheetLike:
        customer = "BLACKWOODS"
        item_code = "LCW77521"
        item = "PB30045075"
        file_stem = "PB30045075 - LCW77521"
        total_rolls = 5.0
        notes = "strong seals no creasing"

    sheet = SheetLike()
    good = score_candidate(sheet, _candidate(item_number="LCW77521"))
    weak = score_candidate(sheet, _candidate(item_number="OTHER-CODE", customer_name="OTHER CUST", quantity=77))
    status, top, _top5 = choose_match([good, weak], high_threshold=90.0, ambiguous_gap=15.0)
    assert status == "matched"
    assert top is not None
    assert top.candidate.myob_item_number == "LCW77521"
    assert top.score > weak.score


def test_primary_match_item_prefix_of_description() -> None:
    class SheetLike:
        customer = "BLACKWOODS"
        item_code = "LCW77521"
        item = "PB30045075"
        file_stem = "PB30045075 - LCW77521"
        total_rolls = 5.0
        notes = None

    cand = CandidateLine(
        order_item_id="oi-2",
        order_id="o-2",
        order_code="INV-002",
        order_date="2026-03-01",
        customer_id="c-1",
        customer_name="BLACKWOODS",
        line_index=0,
        line_kind="myob_import",
        myob_item_number="OTHER",
        import_line_description="PB30045075 LCW77521 STRONG BAG",
        import_ship_quantity=5.0,
        import_quantity_unit="rolls",
        import_qty_type="total_rolls",
        import_requires_job_sheet=True,
        job_sheet_id=None,
    )
    scored = score_candidate(SheetLike(), cand)
    assert "item_matches_description_prefix" in scored.reasons
    assert scored.score >= 130


def test_fallback_item_code_when_item_formula_error() -> None:
    class SheetLike:
        customer = "BLACKWOODS"
        item_code = "LCW77521"
        item = "#VALUE!"
        file_stem = "SOMEFILE"
        total_rolls = None
        notes = None

    cand = CandidateLine(
        order_item_id="oi-3",
        order_id="o-3",
        order_code="INV-003",
        order_date="2026-03-01",
        customer_id="c-1",
        customer_name="BLACKWOODS",
        line_index=0,
        line_kind="myob_import",
        myob_item_number="ZZZ",
        import_line_description="PB30045075 LCW77521 STRONG BAG",
        import_ship_quantity=5.0,
        import_quantity_unit="rolls",
        import_qty_type="total_rolls",
        import_requires_job_sheet=True,
        job_sheet_id=None,
    )
    scored = score_candidate(SheetLike(), cand)
    assert "item_code_matches_description_second_token" in scored.reasons


def test_apply_from_existing_matched_csv_links_job_sheet(tmp_path: Path) -> None:
    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    Sess = sessionmaker(bind=engine)
    db = Sess()
    _seed_myob_draft_placeholders(db)

    cust_id = str(uuid.uuid4())
    order_id = str(uuid.uuid4())
    item_id = str(uuid.uuid4())
    cust = Customer(id=cust_id, name="BLACKWOODS")
    db.add(cust)
    order = Order(id=order_id, code="INV-1", customer_id=cust_id, status=OrderStatus.DISPATCHED)
    db.add(order)
    oi = OrderItem(
        id=item_id,
        order_id=order_id,
        line_index=0,
        line_kind="myob_import",
        import_requires_job_sheet=True,
        import_ship_quantity=5.0,
        import_quantity_unit="rolls",
        import_qty_type="total_rolls",
    )
    db.add(oi)
    db.commit()

    csv_path = tmp_path / "matched.csv"
    csv_path.write_text(
        "order_item_id,import_requires_job_sheet\n"
        f"{item_id},true\n",
        encoding="utf-8",
    )

    out = _apply_from_matched_csv(
        db,
        matched_csv_path=csv_path,
        created_by="test",
        dry_run=False,
    )
    db.commit()
    assert out["rows_in_csv"] == 1
    assert out["created_job_sheets"] == 1
    linked = db.scalar(select(OrderItem).where(OrderItem.id == item_id))
    assert linked is not None
    assert linked.job_sheet_id is not None


def test_apply_from_existing_matched_csv_promotes_placeholder_job_sheet(tmp_path: Path) -> None:
    from app.job_sheets import service as job_sheets_service

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    Sess = sessionmaker(bind=engine)
    db = Sess()
    _seed_myob_draft_placeholders(db)

    cust_id = str(uuid.uuid4())
    order_id = str(uuid.uuid4())
    item_id = str(uuid.uuid4())
    cust = Customer(id=cust_id, name="BLACKWOODS")
    db.add(cust)
    order = Order(id=order_id, code="INV-2", customer_id=cust_id, status=OrderStatus.DISPATCHED)
    db.add(order)
    oi = OrderItem(
        id=item_id,
        order_id=order_id,
        line_index=0,
        line_kind="myob_import",
        import_requires_job_sheet=True,
        import_ship_quantity=5.0,
        import_quantity_unit="rolls",
        import_qty_type="total_rolls",
    )
    db.add(oi)
    db.flush()

    draft_js = job_sheets_service.create_myob_import_draft_job_sheet(
        db=db,
        customer_id=cust_id,
        quantity_value=5.0,
        quantity_unit="rolls",
        qty_type="total_rolls",
        unit_rate=None,
        line_total=None,
        created_by="test",
    )
    oi.job_sheet_id = str(draft_js.id)
    db.add(oi)
    db.commit()

    csv_path = tmp_path / "matched.csv"
    csv_path.write_text(
        "order_item_id,import_requires_job_sheet,item,item_code\n"
        f"{item_id},true,PB30045075,LCW77521\n",
        encoding="utf-8",
    )

    out = _apply_from_matched_csv(
        db,
        matched_csv_path=csv_path,
        created_by="test",
        dry_run=False,
    )
    db.commit()
    assert out["rows_in_csv"] == 1
    assert out["promoted_placeholder_job_sheets"] == 1
    linked = db.scalar(select(OrderItem).where(OrderItem.id == item_id))
    assert linked is not None and linked.job_sheet_id is not None
    js = db.get(JobSheet, linked.job_sheet_id)
    assert js is not None
    assert str(js.product_id) != str(MYOB_DRAFT_PLACEHOLDER_PRODUCT_ID)
    assert bool(js.is_import_draft) is False


def test_apply_from_existing_matched_csv_promotes_duplicates_with_fallback_codes(tmp_path: Path) -> None:
    from app.job_sheets import service as job_sheets_service

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(engine)
    Sess = sessionmaker(bind=engine)
    db = Sess()
    _seed_myob_draft_placeholders(db)

    cust_id = str(uuid.uuid4())
    order_id = str(uuid.uuid4())
    cust = Customer(id=cust_id, name="BLACKWOODS")
    db.add(cust)
    order = Order(id=order_id, code="INV-3", customer_id=cust_id, status=OrderStatus.DISPATCHED)
    db.add(order)
    ids: list[str] = []
    for idx in range(2):
        item_id = str(uuid.uuid4())
        ids.append(item_id)
        oi = OrderItem(
            id=item_id,
            order_id=order_id,
            line_index=idx,
            line_kind="myob_import",
            import_requires_job_sheet=True,
            import_ship_quantity=5.0,
            import_quantity_unit="rolls",
            import_qty_type="total_rolls",
        )
        db.add(oi)
        db.flush()
        js = job_sheets_service.create_myob_import_draft_job_sheet(
            db=db,
            customer_id=cust_id,
            quantity_value=5.0,
            quantity_unit="rolls",
            qty_type="total_rolls",
            unit_rate=None,
            line_total=None,
            created_by="test",
        )
        oi.job_sheet_id = str(js.id)
        db.add(oi)
    db.commit()

    csv_path = tmp_path / "matched.csv"
    csv_path.write_text(
        "order_item_id,import_requires_job_sheet,item,item_code\n"
        f"{ids[0]},true,PB30045075,LCW77521\n"
        f"{ids[1]},true,PB30045075,LCW77521\n",
        encoding="utf-8",
    )

    out = _apply_from_matched_csv(
        db,
        matched_csv_path=csv_path,
        created_by="test",
        dry_run=False,
    )
    db.commit()
    assert out["rows_in_csv"] == 2
    assert out["promoted_placeholder_job_sheets"] == 2
    rows = list(db.scalars(select(OrderItem).where(OrderItem.id.in_(ids))))
    assert len(rows) == 2
    js_rows = [db.get(JobSheet, r.job_sheet_id) for r in rows]
    assert all(js is not None for js in js_rows)
    assert all(str(js.product_id) != str(MYOB_DRAFT_PLACEHOLDER_PRODUCT_ID) for js in js_rows if js is not None)
